import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

class ExcelExporter:
    """Generate Excel files from backtest results"""
    
    @staticmethod
    def create_excel_report(trades_data: list, summary_data: dict) -> BytesIO:
        """
        Create comprehensive Excel report with multiple sheets
        Returns BytesIO object ready for download
        """
        try:
            # Create Excel workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create Summary sheet
            ExcelExporter._create_summary_sheet(wb, summary_data)
            
            # Create Trades sheet
            ExcelExporter._create_trades_sheet(wb, trades_data)
            
            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            return excel_file
            
        except Exception as e:
            logger.error(f"Error creating Excel report: {e}")
            raise
    
    @staticmethod
    def _create_summary_sheet(wb: Workbook, summary: dict):
        """Create summary sheet with key metrics"""
        ws = wb.create_sheet("Summary", 0)
        
        # Header styling
        header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Title
        ws['A1'] = "Backtest Summary Report"
        ws['A1'].font = Font(bold=True, size=16, color="FF6B6B")
        ws.merge_cells('A1:B1')
        
        # Metrics
        metrics = [
            ('', ''),
            ('Strategy', summary.get('strategy', 'N/A')),
            ('Period', summary.get('period', 'N/A')),
            ('Stocks Analyzed', summary.get('stocks_analyzed', 0)),
            ('', ''),
            ('Total Patterns Found', summary.get('total_patterns', 0)),
            ('Target Hit', summary.get('target_hit_count', 0)),
            ('Stop Loss Hit', summary.get('stop_loss_count', 0)),
            ('End of Day Exits', summary.get('eod_exit_count', 0)),
            ('', ''),
            ('Target Hit Rate', f"{summary.get('target_hit_rate', 0):.2f}%"),
            ('Stop Loss Rate', f"{summary.get('stop_loss_rate', 0):.2f}%"),
            ('Average Return', f"{summary.get('avg_return', 0):.2f}%"),
            ('Total Points Gained', f"₹{summary.get('total_points_gained', 0):.2f}"),
        ]
        
        row = 3
        for label, value in metrics:
            if label:  # Skip empty rows
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    @staticmethod
    def _create_trades_sheet(wb: Workbook, trades: list):
        """Create detailed trades sheet"""
        ws = wb.create_sheet("Detailed Trades", 1)
        
        # Headers
        headers = [
            'Stock Symbol', 'Pattern Date', 'Pattern Time', 'Entry Price',
            'Target Price', 'Stop Loss Price', 'Exit Price', 'Exit Time',
            'Exit Reason', 'Points Gained', 'Percentage Return', 
            'Minutes Held', 'Candles Held', 'Outcome'
        ]
        
        # Style for headers
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write trade data
        for row_idx, trade in enumerate(trades, 2):
            ws.cell(row=row_idx, column=1, value=trade.get('stock', ''))
            ws.cell(row=row_idx, column=2, value=trade.get('pattern_date', ''))
            ws.cell(row=row_idx, column=3, value=trade.get('pattern_time', ''))
            ws.cell(row=row_idx, column=4, value=trade.get('entry_price', 0))
            ws.cell(row=row_idx, column=5, value=trade.get('target_price', 0))
            ws.cell(row=row_idx, column=6, value=trade.get('stop_loss_price', 0))
            ws.cell(row=row_idx, column=7, value=trade.get('exit_price', 0))
            ws.cell(row=row_idx, column=8, value=trade.get('exit_time', ''))
            ws.cell(row=row_idx, column=9, value=trade.get('exit_reason', ''))
            ws.cell(row=row_idx, column=10, value=trade.get('points_gained', 0))
            ws.cell(row=row_idx, column=11, value=trade.get('percentage_return', 0))
            ws.cell(row=row_idx, column=12, value=trade.get('minutes_held', 0))
            ws.cell(row=row_idx, column=13, value=trade.get('candles_held', 0))
            ws.cell(row=row_idx, column=14, value=trade.get('outcome', ''))
            
            # Color code outcomes
            outcome_cell = ws.cell(row=row_idx, column=14)
            outcome = trade.get('outcome', '')
            if outcome == 'target_hit':
                outcome_cell.fill = PatternFill(start_color="51CF66", end_color="51CF66", fill_type="solid")
            elif outcome == 'stop_loss':
                outcome_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            else:
                outcome_cell.fill = PatternFill(start_color="FFD43B", end_color="FFD43B", fill_type="solid")
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 15
        
        # Freeze first row
        ws.freeze_panes = 'A2'